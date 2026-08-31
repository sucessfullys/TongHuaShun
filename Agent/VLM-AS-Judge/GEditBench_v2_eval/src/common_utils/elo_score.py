import json
import os
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Sequence

import numpy as np
import pandas as pd
from scipy.optimize import minimize
from scipy.special import logsumexp


@dataclass(frozen=True)
class AggregatedMatches:
    """Sufficient statistics for tie-aware paired comparison on ordered pairs."""

    idx_a: np.ndarray
    idx_b: np.ndarray
    wins_a: np.ndarray
    wins_b: np.ndarray
    ties: np.ndarray
    num_matches: int


@dataclass(frozen=True)
class DimensionData:
    """Parsed data for one evaluation dimension."""

    name: str
    aggregated: AggregatedMatches
    clusters: Sequence[AggregatedMatches]
    observed_models: np.ndarray


def parse_paths(path_str: str) -> List[str]:
    """
    Parse a comma-separated string into a list of file paths.

    Args:
        path_str (str): Input string like "path1,path2,path3"

    Returns:
        List[str]: Parsed list of paths
    """
    if not path_str:
        return []

    return [p.strip() for p in path_str.split(",") if p.strip()]


def parse_csv_values(value_str: Optional[str]) -> List[str]:
    """Parse a comma-separated string into a stripped list of values."""
    if not value_str:
        return []

    return [value.strip() for value in value_str.split(",") if value.strip()]


def _parse_match_key(key: str) -> tuple[str, str, str]:
    """
    Parse a key like "<prompt>_pair_<model_a>_vs_<model_b>".

    Returns:
        tuple[str, str, str]: (prompt_cluster_id, model_a, model_b)
    """
    prompt_cluster_id, pair_part = key.split("_pair_", 1)
    model_a, model_b = pair_part.split("_vs_", 1)
    return prompt_cluster_id, model_a, model_b


def _parse_outcome(winner_str: str) -> float:
    if winner_str == "Image A":
        return 1.0
    if winner_str == "Image B":
        return 0.0
    if winner_str == "Tie":
        return 0.5
    raise ValueError(f"Unsupported winner value: {winner_str}")


def extract_models(
    data_list, excluded_models: Optional[set[str]] = None
) -> set[str]:
    """Extract all model names that appear in a single evaluation dimension."""
    excluded_models = excluded_models or set()
    models = set()
    for item in data_list:
        try:
            _, model_a, model_b = _parse_match_key(item["key"])
        except Exception:
            continue
        if model_a == model_b:
            continue
        if model_a in excluded_models or model_b in excluded_models:
            continue
        models.add(model_a)
        models.add(model_b)
    return models


def parse_data_to_idx(
    data_list, model_to_idx, excluded_models: Optional[set[str]] = None
):
    """
    Parse raw records into an index matrix [idx_a, idx_b, outcome].

    Outcome encoding:
        1.0: Image A wins
        0.0: Image B wins
        0.5: Tie
    """
    excluded_models = excluded_models or set()
    matches = []
    for item in data_list:
        try:
            _, model_a, model_b = _parse_match_key(item["key"])
            winner_str = (item.get("results") or item.get("result") or {})["winner"]
            outcome = _parse_outcome(winner_str)
        except Exception:
            continue

        if model_a == model_b:
            continue
        if model_a in excluded_models or model_b in excluded_models:
            continue
        matches.append([model_to_idx[model_a], model_to_idx[model_b], outcome])

    if not matches:
        return np.empty((0, 3), dtype=float)
    return np.asarray(matches, dtype=float)


def _empty_aggregated_matches() -> AggregatedMatches:
    empty_int = np.empty(0, dtype=np.intp)
    empty_float = np.empty(0, dtype=float)
    return AggregatedMatches(
        idx_a=empty_int,
        idx_b=empty_int,
        wins_a=empty_float,
        wins_b=empty_float,
        ties=empty_float,
        num_matches=0,
    )


def _aggregate_raw_matches(matches_idx: np.ndarray) -> AggregatedMatches:
    """Aggregate repeated ordered pairs into tie-aware sufficient statistics."""
    if len(matches_idx) == 0:
        return _empty_aggregated_matches()

    aggregated = {}
    for idx_a, idx_b, outcome in matches_idx:
        key = (int(idx_a), int(idx_b))
        if key not in aggregated:
            aggregated[key] = [0.0, 0.0, 0.0]
        if outcome == 1.0:
            aggregated[key][0] += 1.0
        elif outcome == 0.0:
            aggregated[key][1] += 1.0
        elif outcome == 0.5:
            aggregated[key][2] += 1.0
        else:
            raise ValueError(f"Unsupported outcome value: {outcome}")

    idx_a = []
    idx_b = []
    wins_a = []
    wins_b = []
    ties = []
    for (a, b), (a_win, b_win, tie_count) in aggregated.items():
        idx_a.append(a)
        idx_b.append(b)
        wins_a.append(a_win)
        wins_b.append(b_win)
        ties.append(tie_count)

    wins_a = np.asarray(wins_a, dtype=float)
    wins_b = np.asarray(wins_b, dtype=float)
    ties = np.asarray(ties, dtype=float)

    return AggregatedMatches(
        idx_a=np.asarray(idx_a, dtype=np.intp),
        idx_b=np.asarray(idx_b, dtype=np.intp),
        wins_a=wins_a,
        wins_b=wins_b,
        ties=ties,
        num_matches=int(round(float(np.sum(wins_a + wins_b + ties)))),
    )


def _combine_cluster_statistics(
    cluster_stats: Sequence[AggregatedMatches], num_models: int
) -> AggregatedMatches:
    """Combine multiple cluster-level sufficient-stat arrays into one dimension."""
    if not cluster_stats:
        return _empty_aggregated_matches()

    wins_a_mat = np.zeros((num_models, num_models), dtype=float)
    wins_b_mat = np.zeros((num_models, num_models), dtype=float)
    ties_mat = np.zeros((num_models, num_models), dtype=float)

    for stats in cluster_stats:
        if stats.num_matches == 0:
            continue
        np.add.at(wins_a_mat, (stats.idx_a, stats.idx_b), stats.wins_a)
        np.add.at(wins_b_mat, (stats.idx_a, stats.idx_b), stats.wins_b)
        np.add.at(ties_mat, (stats.idx_a, stats.idx_b), stats.ties)

    nonzero = (wins_a_mat + wins_b_mat + ties_mat) > 0
    idx_a, idx_b = np.nonzero(nonzero)
    if len(idx_a) == 0:
        return _empty_aggregated_matches()

    wins_a = wins_a_mat[idx_a, idx_b]
    wins_b = wins_b_mat[idx_a, idx_b]
    ties = ties_mat[idx_a, idx_b]
    return AggregatedMatches(
        idx_a=idx_a.astype(np.intp),
        idx_b=idx_b.astype(np.intp),
        wins_a=wins_a,
        wins_b=wins_b,
        ties=ties,
        num_matches=int(round(float((wins_a + wins_b + ties).sum()))),
    )


def _build_dimension_data(
    data_list,
    model_to_idx,
    name: str,
    excluded_models: Optional[set[str]] = None,
) -> DimensionData:
    """
    Parse one dimension into prompt clusters.

    Cluster id is the prompt prefix before "_pair_". Bootstrapping resamples
    these clusters to preserve the correlation induced by reusing the same prompt.
    """
    excluded_models = excluded_models or set()
    cluster_rows: dict[str, list[list[float]]] = {}
    observed_models = set()

    for item in data_list:
        try:
            cluster_id, model_a, model_b = _parse_match_key(item["key"])
            winner_str = (item.get("results") or item.get("result") or {})["winner"]
            outcome = _parse_outcome(winner_str)
        except Exception:
            continue

        if model_a == model_b:
            continue
        if model_a in excluded_models or model_b in excluded_models:
            continue

        idx_a = model_to_idx[model_a]
        idx_b = model_to_idx[model_b]
        cluster_rows.setdefault(cluster_id, []).append([idx_a, idx_b, outcome])
        observed_models.add(idx_a)
        observed_models.add(idx_b)

    cluster_stats = []
    for rows in cluster_rows.values():
        cluster_array = np.asarray(rows, dtype=float)
        cluster_stats.append(_aggregate_raw_matches(cluster_array))

    aggregated = _combine_cluster_statistics(cluster_stats, len(model_to_idx))
    return DimensionData(
        name=name,
        aggregated=aggregated,
        clusters=tuple(cluster_stats),
        observed_models=np.asarray(sorted(observed_models), dtype=np.intp),
    )


def _build_dimension_weights(
    match_stats_list: Sequence[AggregatedMatches], weighting: str
) -> np.ndarray:
    """
    Build dimension-level weights for the joint likelihood.

    balanced:
        Each dimension contributes roughly the same total weight by scaling
        inversely with its number of matches. The overall loss magnitude stays
        close to the unweighted case by using the mean match count as reference.

    by_matches:
        Every pairwise match is counted equally across all dimensions.
    """
    if weighting == "by_matches":
        return np.ones(len(match_stats_list), dtype=float)

    if weighting != "balanced":
        raise ValueError(f"Unsupported dimension weighting mode: {weighting}")

    match_counts = np.asarray(
        [stats.num_matches for stats in match_stats_list], dtype=float
    )
    positive_counts = match_counts[match_counts > 0]
    if len(positive_counts) == 0:
        return np.zeros(len(match_stats_list), dtype=float)

    reference_count = float(np.mean(positive_counts))
    weights = np.zeros(len(match_stats_list), dtype=float)
    for i, match_count in enumerate(match_counts):
        if match_count > 0:
            weights[i] = reference_count / match_count
    return weights


def fit_bradley_terry(matches_idx, num_models, alpha=1):
    """
    Fit a single-dimension paired-comparison model by maximum likelihood.

    This legacy entry point is retained for compatibility and internally reuses
    the single-dimension path of the joint optimizer.
    If ties are present, the Davidson extension is used for explicit tie modeling.
    """
    if num_models <= 1:
        return np.zeros(num_models)

    aggregated = _aggregate_raw_matches(np.asarray(matches_idx, dtype=float))
    theta, _ = _fit_joint_paired_comparison_model(
        [aggregated],
        num_models=num_models,
        alpha=float(alpha),
        dimension_weights=np.ones(1, dtype=float),
    )
    return theta


def _fit_joint_paired_comparison_model(
    match_stats_list: Sequence[AggregatedMatches],
    num_models: int,
    alpha: float = 1.0,
    dimension_weights: Optional[np.ndarray] = None,
    initial_theta: Optional[np.ndarray] = None,
    initial_log_nu: Optional[np.ndarray] = None,
    tie_log_nu_bounds: tuple[float, float] = (-12.0, 8.0),
) -> tuple[np.ndarray, np.ndarray]:
    """
    Fit one shared ability vector across all observed dimensions.

    If ties are present, this uses the Davidson extension:
        P(A)   = exp(theta_a) / Z
        P(B)   = exp(theta_b) / Z
        P(Tie) = 2 * nu_d * exp((theta_a + theta_b) / 2) / Z
    where nu_d is a dimension-specific tie propensity and
        Z = exp(theta_a) + exp(theta_b) + 2 * nu_d * exp((theta_a + theta_b) / 2)

    Using one shared theta across all dimensions avoids assigning fake
    zero-centered scores to models that are missing from some dimensions,
    while the per-dimension tie parameter lets each metric have its own tie rate.
    """
    if num_models == 0:
        return np.empty(0, dtype=float), np.empty(len(match_stats_list), dtype=float)
    if num_models == 1:
        return np.zeros(1, dtype=float), np.zeros(len(match_stats_list), dtype=float)

    if dimension_weights is None:
        dimension_weights = np.ones(len(match_stats_list), dtype=float)
    else:
        dimension_weights = np.asarray(dimension_weights, dtype=float)

    if initial_theta is None:
        free_initial_theta = np.zeros(num_models - 1, dtype=float)
    else:
        initial_theta = np.asarray(initial_theta, dtype=float)
        if initial_theta.shape != (num_models,):
            raise ValueError("initial_theta shape does not match num_models")
        centered_theta = initial_theta - np.mean(initial_theta)
        free_initial_theta = centered_theta[:-1]

    if initial_log_nu is None:
        initial_log_nu = []
        for stats in match_stats_list:
            tie_rate = 0.0 if stats.num_matches == 0 else float(stats.ties.sum()) / stats.num_matches
            tie_rate = float(np.clip(tie_rate, 1e-6, 1.0 - 1e-6))
            initial_log_nu.append(np.log(tie_rate / (1.0 - tie_rate)))
        initial_log_nu = np.asarray(initial_log_nu, dtype=float)
    else:
        initial_log_nu = np.asarray(initial_log_nu, dtype=float)
        if initial_log_nu.shape != (len(match_stats_list),):
            raise ValueError("initial_log_nu shape does not match number of dimensions")

    lower_log_nu, upper_log_nu = tie_log_nu_bounds
    initial_log_nu = np.clip(initial_log_nu, lower_log_nu, upper_log_nu)
    initial_params = np.concatenate([free_initial_theta, initial_log_nu])
    bounds = [(None, None)] * (num_models - 1) + [tie_log_nu_bounds] * len(match_stats_list)

    def objective(params):
        free_theta = params[: num_models - 1]
        log_nu = params[num_models - 1 :]
        theta = np.concatenate([free_theta, np.array([0.0])])

        loss = 0.5 * alpha * np.dot(theta, theta)
        grad = alpha * theta
        grad_log_nu = np.zeros(len(match_stats_list), dtype=float)

        for dim_idx, (weight, stats) in enumerate(zip(dimension_weights, match_stats_list)):
            if weight == 0.0 or stats.num_matches == 0:
                continue

            theta_a = theta[stats.idx_a]
            theta_b = theta[stats.idx_b]
            tie_logit = log_nu[dim_idx] + 0.5 * (theta_a + theta_b) + np.log(2.0)
            logits = np.column_stack([theta_a, theta_b, tie_logit])
            log_z = logsumexp(logits, axis=1)

            log_p_a = theta_a - log_z
            log_p_b = theta_b - log_z
            log_p_tie = tie_logit - log_z
            total_count = stats.wins_a + stats.wins_b + stats.ties

            loss -= weight * np.sum(
                stats.wins_a * log_p_a
                + stats.wins_b * log_p_b
                + stats.ties * log_p_tie
            )

            probs = np.exp(logits - log_z[:, None])
            grad_a_logit = weight * (total_count * probs[:, 0] - stats.wins_a)
            grad_b_logit = weight * (total_count * probs[:, 1] - stats.wins_b)
            grad_tie_logit = weight * (total_count * probs[:, 2] - stats.ties)

            np.add.at(grad, stats.idx_a, grad_a_logit + 0.5 * grad_tie_logit)
            np.add.at(grad, stats.idx_b, grad_b_logit + 0.5 * grad_tie_logit)
            grad_log_nu[dim_idx] += np.sum(grad_tie_logit)

        return loss, np.concatenate([grad[:-1], grad_log_nu])

    res = minimize(
        objective,
        initial_params,
        method="L-BFGS-B",
        jac=True,
        bounds=bounds,
    )
    if not res.success:
        raise RuntimeError(f"Paired comparison optimization failed: {res.message}")

    theta = np.concatenate([res.x[: num_models - 1], np.array([0.0])])
    theta -= np.mean(theta)
    log_nu = res.x[num_models - 1 :]
    return theta, log_nu


def _resample_dimension_clusters(
    dimension_data: DimensionData, num_models: int, rng: np.random.Generator
) -> AggregatedMatches:
    """Bootstrap one dimension by resampling prompt clusters with replacement."""
    num_clusters = len(dimension_data.clusters)
    if num_clusters == 0:
        return _empty_aggregated_matches()

    sampled_indices = rng.integers(0, num_clusters, size=num_clusters)
    sampled_clusters = [dimension_data.clusters[i] for i in sampled_indices]
    return _combine_cluster_statistics(sampled_clusters, num_models)


def _connected_component_sizes(
    match_stats_list: Sequence[AggregatedMatches], num_models: int
) -> list[int]:
    """Return connected-component sizes for the overall comparison graph."""
    if num_models == 0:
        return []

    parent = list(range(num_models))
    observed = np.zeros(num_models, dtype=bool)

    def find(x: int) -> int:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a: int, b: int) -> None:
        root_a = find(a)
        root_b = find(b)
        if root_a != root_b:
            parent[root_b] = root_a

    for stats in match_stats_list:
        for idx_a, idx_b in zip(stats.idx_a, stats.idx_b):
            observed[idx_a] = True
            observed[idx_b] = True
            union(int(idx_a), int(idx_b))

    component_sizes = {}
    for idx, is_observed in enumerate(observed):
        if not is_observed:
            continue
        root = find(idx)
        component_sizes[root] = component_sizes.get(root, 0) + 1
    return sorted(component_sizes.values(), reverse=True)


def calculate_joint_leaderboard(
    dim_data_list,
    n_bootstrap: int = 100,
    dimension_names: Optional[Sequence[str]] = None,
    alpha: float = 1.0,
    dimension_weighting: str = "balanced",
    random_seed: Optional[int] = None,
    exclude_models: Optional[Sequence[str]] = None,
):
    """
    Compute the joint leaderboard and estimate confidence intervals using
    prompt-cluster bootstrap.

    Key modeling choices:
    1. A shared theta is fit directly on the joint likelihood across dimensions,
       rather than fitting each dimension independently and averaging scores.
    2. Bootstrap resampling is performed at the prompt-cluster level so that
       multiple pairwise matches derived from the same prompt are not treated as
       independent samples.
    3. Ties are modeled explicitly with the Davidson extension, with one tie
       propensity parameter per dimension.
    """
    excluded_models = set(exclude_models or [])
    all_models = set()
    for data in dim_data_list:
        all_models.update(extract_models(data, excluded_models=excluded_models))

    model_names = sorted(all_models)
    num_models = len(model_names)
    if num_models == 0:
        return pd.DataFrame(
            columns=[
                "Model",
                "Score",
                "CI_String",
                "Dimensions_Covered",
                "Match_Count",
            ]
        )

    model_to_idx = {name: i for i, name in enumerate(model_names)}
    num_dimensions = len(dim_data_list)
    if dimension_names is None:
        dimension_names = [f"dim_{i}" for i in range(num_dimensions)]
    elif len(dimension_names) != num_dimensions:
        raise ValueError("dimension_names length must match dim_data_list length")

    dimensions = [
        _build_dimension_data(
            data, model_to_idx, name, excluded_models=excluded_models
        )
        for data, name in zip(dim_data_list, dimension_names)
    ]
    match_stats_list = [dimension.aggregated for dimension in dimensions]
    dimension_weights = _build_dimension_weights(
        match_stats_list, weighting=dimension_weighting
    )

    component_sizes = _connected_component_sizes(match_stats_list, num_models)
    if len(component_sizes) > 1:
        print(
            "Warning: comparison graph has "
            f"{len(component_sizes)} disconnected components {component_sizes}; "
            "rank gaps across components rely on regularization.",
            file=sys.stderr,
        )

    theta_overall_main, log_nu_main = _fit_joint_paired_comparison_model(
        match_stats_list,
        num_models=num_models,
        alpha=alpha,
        dimension_weights=dimension_weights,
    )

    if n_bootstrap > 0:
        print(
            f"Running {n_bootstrap} joint bootstrap resamples at the prompt-cluster level...",
            flush=True,
        )

    rng = np.random.default_rng(random_seed)
    bootstrap_overall_thetas = []
    bootstrap_failures = 0

    for _ in range(max(n_bootstrap, 0)):
        sampled_match_stats = [
            _resample_dimension_clusters(dimension, num_models, rng)
            for dimension in dimensions
        ]
        sampled_weights = _build_dimension_weights(
            sampled_match_stats, weighting=dimension_weighting
        )

        try:
            theta_sample, _ = _fit_joint_paired_comparison_model(
                sampled_match_stats,
                num_models=num_models,
                alpha=alpha,
                dimension_weights=sampled_weights,
                initial_theta=theta_overall_main,
                initial_log_nu=log_nu_main,
            )
            bootstrap_overall_thetas.append(theta_sample)
        except RuntimeError:
            bootstrap_failures += 1

    if bootstrap_overall_thetas:
        bootstrap_overall_thetas = np.asarray(bootstrap_overall_thetas)
        ci_lower_theta = np.percentile(bootstrap_overall_thetas, 2.5, axis=0)
        ci_upper_theta = np.percentile(bootstrap_overall_thetas, 97.5, axis=0)
    else:
        ci_lower_theta = theta_overall_main.copy()
        ci_upper_theta = theta_overall_main.copy()

    if bootstrap_failures > 0:
        print(
            f"Warning: {bootstrap_failures} bootstrap fits failed and were skipped.",
            file=sys.stderr,
        )

    scale_factor = 400 / np.log(10)
    elo_main = 1000 + theta_overall_main * scale_factor
    elo_lower = 1000 + ci_lower_theta * scale_factor
    elo_upper = 1000 + ci_upper_theta * scale_factor

    coverage_counts = np.zeros(num_models, dtype=int)
    match_counts = np.zeros(num_models, dtype=int)
    for dimension in dimensions:
        coverage_counts[dimension.observed_models] += 1
        total_count = (
            dimension.aggregated.wins_a
            + dimension.aggregated.wins_b
            + dimension.aggregated.ties
        ).astype(int)
        np.add.at(match_counts, dimension.aggregated.idx_a, total_count)
        np.add.at(match_counts, dimension.aggregated.idx_b, total_count)

    results = []
    for i, name in enumerate(model_names):
        score = elo_main[i]
        offset_lower = int(round(elo_lower[i] - score))
        offset_upper = int(round(elo_upper[i] - score))
        ci_str = (
            f"{offset_lower}/+{offset_upper}"
            if offset_upper > 0
            else f"{offset_lower}/{offset_upper}"
        )

        results.append(
            {
                "Model": name,
                "Score": int(round(score)),
                "CI_String": ci_str,
                "Dimensions_Covered": int(coverage_counts[i]),
                "Match_Count": int(match_counts[i]),
            }
        )

    df = pd.DataFrame(results)
    df = df.sort_values(by="Score", ascending=False).reset_index(drop=True)
    df.index = df.index + 1
    df.attrs["dimension_tie_log_nu"] = {
        dimension.name: float(log_nu)
        for dimension, log_nu in zip(dimensions, log_nu_main)
    }
    df.attrs["dimension_tie_nu"] = {
        dimension.name: float(np.exp(log_nu))
        for dimension, log_nu in zip(dimensions, log_nu_main)
    }
    return df


def print_leaderboard(title: str, leaderboard_df: pd.DataFrame) -> None:
    print("\n" + "=" * 55)
    print(title)
    print("-" * 55)
    print(f"{'Rank':<5} | {'Model':<25} | {'ELO ↑':<6} | {'95% CI'}")
    print("-" * 55)
    for rank, row in leaderboard_df.iterrows():
        print(f"{rank:<5} | {row['Model']:<25} | {row['Score']:<6} | {row['CI_String']}")
    print("=" * 55)

    if leaderboard_df.attrs.get("dimension_tie_nu"):
        print("\nTie Propensity By Dimension (Davidson nu)")
        print("-" * 55)
        for dim_name, nu in leaderboard_df.attrs["dimension_tie_nu"].items():
            print(f"{dim_name:<30} | nu={nu:.6f}")


def _infer_dimension_display_name(name: str) -> str:
    """Map result-file names to human-readable column titles."""
    lookup_text = str(name).lower()
    if "eval_if_" in lookup_text or "_if_" in lookup_text:
        return "Instruction Following"
    if "eval_vq_" in lookup_text or "_vq_" in lookup_text:
        return "Visual Quality"
    if "eval_vc_" in lookup_text or "_vc_" in lookup_text:
        return "Visual Consistency"
    return Path(name).stem.replace("_", " ").title()


def _collect_model_sample_counts(
    dim_data_list, excluded_models: Optional[Sequence[str]] = None
) -> dict[str, int]:
    """
    Count unique prompt clusters per model across all provided dimensions.

    This approximates the "Samples" column in benchmark tables by counting how
    many unique prompt ids each model actually appears in.
    """
    excluded_models = set(excluded_models or [])
    model_to_clusters: dict[str, set[str]] = {}

    for data_list in dim_data_list:
        for item in data_list:
            try:
                cluster_id, model_a, model_b = _parse_match_key(item["key"])
            except Exception:
                continue

            if model_a == model_b:
                continue

            for model_name in (model_a, model_b):
                if model_name in excluded_models:
                    continue
                model_to_clusters.setdefault(model_name, set()).add(cluster_id)

    return {
        model_name: len(cluster_ids)
        for model_name, cluster_ids in model_to_clusters.items()
    }


def _build_summary_table(
    overall_df: pd.DataFrame,
    dimension_tables: Sequence[tuple[str, pd.DataFrame]],
    sample_counts: dict[str, int],
) -> pd.DataFrame:
    """Build a wide summary table with grouped headers like the paper figure."""
    model_order = overall_df["Model"].tolist()
    summary_columns = [
        ("Model", ""),
        ("Samples", ""),
    ]

    for display_name, _ in dimension_tables:
        summary_columns.extend(
            [
                (display_name, "ELO↑"),
                (display_name, "95% CI"),
            ]
        )

    summary_columns.extend(
        [
            ("Overall", "ELO↑"),
            ("Overall", "95% CI"),
            ("Rank", ""),
        ]
    )

    rows = []
    dimension_lookup = [
        (display_name, df.set_index("Model"))
        for display_name, df in dimension_tables
    ]
    overall_lookup = overall_df.set_index("Model")

    for rank, model_name in enumerate(model_order, start=1):
        row = {
            ("Model", ""): model_name,
            ("Samples", ""): sample_counts.get(model_name, 0),
        }

        for display_name, dim_df in dimension_lookup:
            if model_name in dim_df.index:
                row[(display_name, "ELO↑")] = int(dim_df.at[model_name, "Score"])
                row[(display_name, "95% CI")] = dim_df.at[model_name, "CI_String"]
            else:
                row[(display_name, "ELO↑")] = ""
                row[(display_name, "95% CI")] = ""

        row[("Overall", "ELO↑")] = int(overall_lookup.at[model_name, "Score"])
        row[("Overall", "95% CI")] = overall_lookup.at[model_name, "CI_String"]
        row[("Rank", "")] = rank
        rows.append(row)

    summary_df = pd.DataFrame(rows)
    summary_df = summary_df.reindex(columns=pd.MultiIndex.from_tuples(summary_columns))
    return summary_df


def _flatten_summary_columns(summary_df: pd.DataFrame) -> pd.DataFrame:
    flat_df = summary_df.copy()
    flat_df.columns = [
        top if not sub else f"{top} {sub}" for top, sub in flat_df.columns.to_list()
    ]
    return flat_df


def _export_summary_table_to_xlsx(
    output_path: str,
    summary_df: pd.DataFrame,
    overall_df: pd.DataFrame,
    dimension_tables: Sequence[tuple[str, pd.DataFrame]],
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError(
            "Exporting .xlsx tables requires openpyxl to be installed."
        ) from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Summary"

    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    subheader_fill = PatternFill("solid", fgColor="EAF1FB")
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    columns = list(summary_df.columns)
    col_idx = 1
    while col_idx <= len(columns):
        top_label, sub_label = columns[col_idx - 1]
        group_start = col_idx
        while col_idx <= len(columns) and columns[col_idx - 1][0] == top_label:
            col_idx += 1
        group_end = col_idx - 1

        top_cell = worksheet.cell(row=1, column=group_start, value=top_label)
        top_cell.font = header_font
        top_cell.fill = header_fill
        top_cell.alignment = align_center
        top_cell.border = border

        if group_start == group_end and sub_label == "":
            worksheet.merge_cells(
                start_row=1,
                start_column=group_start,
                end_row=2,
                end_column=group_end,
            )
        else:
            worksheet.merge_cells(
                start_row=1,
                start_column=group_start,
                end_row=1,
                end_column=group_end,
            )
            for sub_col in range(group_start, group_end + 1):
                sub_cell = worksheet.cell(
                    row=2, column=sub_col, value=columns[sub_col - 1][1]
                )
                sub_cell.font = header_font
                sub_cell.fill = subheader_fill
                sub_cell.alignment = align_center
                sub_cell.border = border

    for row_idx, (_, row) in enumerate(summary_df.iterrows(), start=3):
        for col_idx, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = align_left if col_idx == 1 else align_center

    width_map = {
        "A": 28,
        "B": 10,
    }
    for col_letter, width in width_map.items():
        worksheet.column_dimensions[col_letter].width = width

    for col_idx, (top_label, sub_label) in enumerate(columns[2:], start=3):
        column_letter = worksheet.cell(row=1, column=col_idx).column_letter
        if sub_label == "95% CI":
            worksheet.column_dimensions[column_letter].width = 12
        elif top_label == "Rank":
            worksheet.column_dimensions[column_letter].width = 8
        else:
            worksheet.column_dimensions[column_letter].width = 10

    worksheet.freeze_panes = "A3"

    detail_frames = [("Overall", overall_df)] + list(dimension_tables)
    for sheet_name, detail_df in detail_frames:
        detail_ws = workbook.create_sheet(title=sheet_name[:31])
        export_df = detail_df.copy().reset_index(drop=True)
        export_df.insert(0, "Rank", np.arange(1, len(export_df) + 1))
        for row_idx, row_values in enumerate(
            [export_df.columns.tolist()] + export_df.values.tolist(), start=1
        ):
            for col_idx, value in enumerate(row_values, start=1):
                cell = detail_ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = align_left if col_idx == 2 else align_center
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
        detail_ws.freeze_panes = "A2"

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    workbook.save(output_path)


def _export_summary_table_to_html(
    output_path: str,
    summary_df: pd.DataFrame,
    overall_df: pd.DataFrame,
    dimension_tables: Sequence[tuple[str, pd.DataFrame]],
) -> None:
    summary_html = summary_df.to_html(index=False, border=0, classes=["elo-summary"])

    detail_sections = []
    detail_frames = [("Overall", overall_df)] + list(dimension_tables)
    for section_title, detail_df in detail_frames:
        export_df = detail_df.copy().reset_index(drop=True)
        export_df.insert(0, "Rank", np.arange(1, len(export_df) + 1))
        detail_sections.append(
            f"<h2>{section_title}</h2>\n"
            + export_df.to_html(index=False, border=0, classes=["elo-detail"])
        )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>ELO Leaderboard</title>
  <style>
    body {{
      font-family: Arial, sans-serif;
      margin: 24px;
      color: #222;
    }}
    h1, h2 {{
      margin: 0 0 12px;
    }}
    h2 {{
      margin-top: 28px;
    }}
    table {{
      border-collapse: collapse;
      margin: 12px 0 20px;
      width: 100%;
      font-size: 14px;
    }}
    th, td {{
      border: 1px solid #9aa5b1;
      padding: 6px 10px;
      text-align: center;
      white-space: nowrap;
    }}
    thead th {{
      background: #d9e2f3;
      font-weight: 700;
    }}
    thead tr:nth-child(2) th {{
      background: #eaf1fb;
    }}
    td:first-child, th:first-child {{
      text-align: left;
    }}
  </style>
</head>
<body>
  <h1>ELO Leaderboard Summary</h1>
  {summary_html}
  {''.join(detail_sections)}
</body>
</html>
"""

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)


def export_elo_table(
    output_path: str,
    overall_df: pd.DataFrame,
    dimension_tables: Sequence[tuple[str, pd.DataFrame]],
    sample_counts: dict[str, int],
) -> None:
    summary_df = _build_summary_table(
        overall_df=overall_df,
        dimension_tables=dimension_tables,
        sample_counts=sample_counts,
    )
    suffix = Path(output_path).suffix.lower()

    if suffix == ".xlsx":
        _export_summary_table_to_xlsx(
            output_path=output_path,
            summary_df=summary_df,
            overall_df=overall_df,
            dimension_tables=dimension_tables,
        )
        return

    if suffix == ".html":
        _export_summary_table_to_html(
            output_path=output_path,
            summary_df=summary_df,
            overall_df=overall_df,
            dimension_tables=dimension_tables,
        )
        return

    flat_df = _flatten_summary_columns(summary_df)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    if suffix == ".csv":
        flat_df.to_csv(output_path, index=False, encoding="utf-8-sig")
        return
    raise ValueError("Unsupported table output format. Use .xlsx, .html, or .csv")


def parse_args():
    import argparse

    parser = argparse.ArgumentParser(
        description=(
            "Calculate leaderboard scores and confidence intervals using a "
            "joint Davidson/Bradley-Terry model with prompt-cluster bootstrap."
        )
    )
    parser.add_argument(
        "--result-files",
        type=str,
        required=True,
        help="Comma-separated paths to JSONL files containing pairwise match results.",
    )
    parser.add_argument(
        "--bootstrap",
        type=int,
        default=100,
        help="Number of prompt-cluster bootstrap samples for confidence interval estimation.",
    )
    parser.add_argument(
        "--alpha",
        type=float,
        default=1.0,
        help="L2 regularization coefficient for the shared ability parameters.",
    )
    parser.add_argument(
        "--dimension-weighting",
        type=str,
        default="balanced",
        choices=["balanced", "by_matches"],
        help=(
            "How to weight multiple dimensions in the joint likelihood: "
            "'balanced' equalizes total weight per dimension, "
            "'by_matches' counts every pairwise match equally."
        ),
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=None,
        help="Random seed used for bootstrap resampling.",
    )
    parser.add_argument(
        "--exclude-models",
        type=str,
        default=None,
        help=(
            "Comma-separated model names to remove entirely before fitting. "
            "Any match involving one of these models will be skipped."
        ),
    )
    parser.add_argument(
        "--table-output",
        type=str,
        default=None,
        help=(
            "Optional output path for a formatted leaderboard table. "
            "Supports .xlsx, .html, and .csv."
        ),
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    files = parse_paths(args.result_files)
    exclude_models = parse_csv_values(args.exclude_models)
    dimension_names = files

    dim_data_list = []
    for res_file in files:
        with open(res_file, "r", encoding="utf-8") as f:
            dim_data_list.append([json.loads(line) for line in f])

    overall_leaderboard_df = None
    if len(files) >= 2:
        leaderboard_df = calculate_joint_leaderboard(
            dim_data_list,
            n_bootstrap=args.bootstrap,
            dimension_names=dimension_names,
            alpha=args.alpha,
            dimension_weighting=args.dimension_weighting,
            random_seed=args.seed,
            exclude_models=exclude_models,
        )
        overall_leaderboard_df = leaderboard_df
        print_leaderboard("Overall Joint Leaderboard", leaderboard_df)

    dimension_tables = []
    for idx, (dim_name, dim_data) in enumerate(zip(dimension_names, dim_data_list)):
        dim_seed = None if args.seed is None else args.seed + idx + 1
        dimension_df = calculate_joint_leaderboard(
            [dim_data],
            n_bootstrap=args.bootstrap,
            dimension_names=[dim_name],
            alpha=args.alpha,
            dimension_weighting=args.dimension_weighting,
            random_seed=dim_seed,
            exclude_models=exclude_models,
        )
        dimension_tables.append((_infer_dimension_display_name(dim_name), dimension_df))
        print(f"\nResult file: {files[idx]}")
        print_leaderboard(f"Single-Dimension Leaderboard: {dim_name}", dimension_df)

    if args.table_output:
        if overall_leaderboard_df is None:
            overall_leaderboard_df = dimension_tables[0][1]
        sample_counts = _collect_model_sample_counts(
            dim_data_list, excluded_models=exclude_models
        )
        export_elo_table(
            output_path=args.table_output,
            overall_df=overall_leaderboard_df,
            dimension_tables=dimension_tables,
            sample_counts=sample_counts,
        )
        print(f"\nSaved leaderboard table to: {args.table_output}")
 